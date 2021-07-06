from ast import Index
from os import truncate
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from unidecode import unidecode

#
# This gets you 9/10 of the way there. To finish up, make a new column with the list from procyclingstats sorted alphabetically and compare it to the output of
# chosenOnes to see which names were entered differently in the pcm game (i.e. Sergio Henao vs Sergio Luis Henao)
#
def CullCyclists():
    print("Culling from AllCyclists.xlxs to SelectedRiders.xlxs")

    workbook = load_workbook(filename="SelectedRiders.xlsx")
    #print(workbook.sheetnames)
    sheet = workbook["DYN_cyclist"]


    namesText = open("SelectedRiders.txt", "r")
    textLines = namesText.read().splitlines()

    rows, cols = (2,len(textLines))
    names = [[0 for i in range(cols)] for j in range(rows)]


    for i in range(cols):
        splitLine = textLines[i].split()
        #print(splitLine)
        numberOfStrings = len(splitLine)
        firstName = ""
        lastName = ""
        #print(firstName + " " + str(numberOfSplits))
        count = 0
        while count < numberOfStrings:
            if unidecode(splitLine[count]).upper() == unidecode(splitLine[count]):
                # it is all caps, therefore lastName
                if lastName == "":
                    lastName += unidecode(splitLine[count]).lower()
                else:
                    lastName += " "
                    lastName += unidecode(splitLine[count]).lower()
            else:
                if firstName == "":
                    firstName += unidecode(splitLine[count]).lower()
                else:
                    firstName += " "
                    firstName += unidecode(splitLine[count]).lower()
            count += 1

        # print(firstName + " ||| " + lastName)
        names[0][i] = firstName
        names[1][i] = lastName

    deletionRows = []
    selectionCount = 0
    #print ("rows " + str(sheet.max_row))
    for i in range(2, sheet.max_row + 1):
        #print(sheet.cell(i,2).value)
        #print(type(sheet.cell(i,2).value))

        lastName = unidecode(sheet.cell(i,2).value).lower()
        firstName = unidecode(sheet.cell(i,3).value).lower()
        #print(str(i) + " Testing rider, sheet: " + firstName +" "+ lastName)


        if (names[0].__contains__(firstName)):
            matchingFirstNames = [q for q, x in enumerate(names[0]) if x == firstName]
            foundMatch = False

            for indix in matchingFirstNames:
                #print(str(indix) + " " + names[1][indix])
                if names[1][indix] == lastName:
                    foundMatch = True
                    #print("Last Name found at: " + str(indix) + ", " + names[1][indix])

            if foundMatch:
                selectionCount+=1
                #print(str(selectionCount)+ " SELECTED " + firstName + " " + lastName)

            else:
                #print (" Second name no match: " + firstName + " " + lastName)
                deletionRows.append(i)

        else:
            #print("First name not found: " + firstName + " " + lastName)
            deletionRows.append(i)

    tuples = reverseCombiner(deletionRows)
    for tuple in tuples:
        sheet.delete_rows(tuple[0], tuple[1])
    workbook.save('SelectedRiders.xlsx')
    
def reverseCombiner(rowList):
    # Don't do anything for empty list. Otherwise,
    # make a copy and sort.

    if len(rowList) == 0: return []
    sortedList = rowList[:]
    sortedList.sort()

    # Init, empty tuple, use first item for previous and
    # first in this run.

    tupleList = []
    firstItem = sortedList[0]
    prevItem = sortedList[0]

    # Process all other items in order.

    for item in sortedList[1:]:
        # If start of new run, add tuple and use new first-in-run.

        if item != prevItem + 1:
            tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
            firstItem = item

        # Regardless, current becomes previous for next loop.

        prevItem = item

    # Finish off the final run and return tuple list.

    tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
    return tupleList




if __name__ == '__main__':
    CullCyclists()


# Check for completion, mismatches in the db
