from Rider import Rider
from openpyxl import load_workbook
from unidecode import unidecode
from openpyxl.worksheet import worksheet
import time
from decimal import *

selectedRidersWorkbook = load_workbook(filename="SelectedRiders.xlsx")
sheet = selectedRidersWorkbook["DYN_cyclist"]
pointsSheet = load_workbook(filename="VelogamesPoints.xlsx")["PointsSheet"]
riders = []

def LoadStats():
    for i in range(184):
        r = Rider()
        nam = str(sheet.cell(i+2,3).value) + " " + str(sheet.cell(i+2,2).value)
        nam = unidecode(nam)
        r.name = nam

        r.flat= sheet.cell(i+2,34).value
        r.mountain= sheet.cell(i+2,36).value
        r.downhill = sheet.cell(i+2, 38).value
        r.cobbles = sheet.cell(i+2, 40).value
        r.tt = sheet.cell(i+2,42).value
        r.prologue = sheet.cell(i+2, 44).value
        r.sprint= sheet.cell(i+2,46).value
        r.acceleration= sheet.cell(i+2,48).value
        r.endurance = sheet.cell(i+2, 50).value
        r.resistance= sheet.cell(i+2,52).value
        r.recovery = sheet.cell(i+2, 54).value
        r.hill= sheet.cell(i+2,56).value
        r.index = i + 2
        riders.append(r)
        #print(str(i) + " " + nam)
        #if i == 33:
            #print(r.name + " {0} {1} {2} {4} {5} {6} {7} {8} {9} ".format(r.flat, r.mountain, r.downhill, r.tt, r.sprint, r.acceleration,r.resistance, r.recovery,r.hill, r.index))
def RecoveryFactor(rider, StageNumber):
        return (rider.recovery + (100 - rider.recovery) * (1 - StageNumber/21/3))

def SprintStage(StageNumber):
    # 100% of sprint at stage 1, recovery% at stage 21          65       + 35 * 20/21
    riders.sort(key=lambda a: (3 * a.sprint + a.acceleration) * RecoveryFactor(a,StageNumber), reverse=True)
    #add points to the riders
    PrintStageResults("Sprint", StageNumber)

    riders[1].timegap += 4
    riders[2].timegap += 6

    for i in range(3, len(riders)):
        riders[i].timegap +=(10 + 0.001 * i)

    GiveWinPoints(StageNumber)
    AssignGCPoints(StageNumber)


def HillStage(StageNumber):

    riders.sort(key=lambda a: (2*a.hill + a.acceleration/2 + a.resistance/2) * RecoveryFactor(a,StageNumber), reverse=True)
    #add points to the riders
    PrintStageResults("Hilly", StageNumber)
    GiveWinPoints(StageNumber)
    for i in range(10):
        #timegaps. add 1 sec for each top 10, then 5 sec for next 10 each and so on
        riders[i].timegap += i
        #print(riders[i].name + " " +str( riders[i].timegap))
    for i in range(10, len(riders)):
        riders[i].timegap += 10 + int(i/10) * 5
        #print(riders[i].name + " " + str(riders[i].timegap))
    AssignGCPoints(StageNumber)

def DownHill(StageNumber):
    riders.sort(key=lambda a: 2*a.mountain + a.downhill/4 + a.endurance/6, reverse = True)
    selection = []
    for i in range(20):
        selection.append(riders[i])

    selection.sort(key=lambda a: (2 * a.mountain + a.sprint + a.downhill) * RecoveryFactor(a,StageNumber), reverse = True)
    for i in range(20):
        riders[i] = selection[i]

    PrintStageResults("DownHill mountain finish", StageNumber)
    GiveWinPoints(StageNumber)

    for i in range(20):
        #timegaps. add 1 sec for each top 10, then 5 sec for next 10 each and so on
        riders[i].timegap += i
        #print(riders[i].name + " " +str( riders[i].timegap))
    for i in range(20, len(riders)):
        riders[i].timegap += (20 + int(i/20) * 60)
        #print(riders[i].name + " " + str(riders[i].timegap))
    AssignGCPoints(StageNumber)

def MTF(StageNumber):
    riders.sort(key=lambda a: (2* a.mountain + a.hill/8 + a.endurance/6) * RecoveryFactor(a,StageNumber), reverse=True)
    PrintStageResults('MTF', StageNumber)
    GiveWinPoints(StageNumber)
    for i in range(len(riders)):
        riders[i].timegap += i * 10
    AssignGCPoints(StageNumber)

def HillyITT(StageNumber):
    riders.sort(key=lambda a: (a.flat + a.hill + 5*a.tt + a.resistance) * RecoveryFactor(a,StageNumber), reverse=True)
    PrintStageResults("Hilly ITT", StageNumber)
    GiveWinPoints(StageNumber)
    for i in range(len(riders)):
        riders[i].timegap += i * 4
    AssignGCPoints(StageNumber)

def AssignGCPoints(StageNumber):
    riders.sort(key=lambda a: a.timegap)
    #print("Giving points: Stage " + str(StageNumber))
    for i in range(20):
        GCPoints = 0
        for q in range(7):
            GCPoints += pointsSheet.cell(2 + q + i, 5).value
        GCPoints/=7
        #print("Giving " + "%0.2f" % (GCPoints,) + "GC points to " + riders[i].name)
        riders[i].totalPoints += GCPoints
    PrintGC(StageNumber)

def GiveWinPoints(StageNumber):
    for i, rider in enumerate(riders):
        if i < 20:
            winPoints = 0
            for q in range(7):
                winPoints += pointsSheet.cell(2 + q + i, 2).value
            winPoints/=7
            #print("Giving " + "%0.2f" % (winPoints,) + " WIN points to " + rider.name)

            rider.totalPoints += winPoints
            #print(rider.name + str(rider.totalPoints))

def PrintStageResults(stageType, stageNumber):
    print("Stage ", str(stageNumber), " Stage Results, ", stageType)
    for i in range(10):
        print(str(i+1), ": " + riders[i].name)
    pass

def PrintGC(stageNumber):
    print("Stage " + str(stageNumber) + " GC")
    leadersTime = riders[0].timegap
    for i in range(10):
        print(str(i + 1) + " :" + riders[i].name + " + " + str(time.strftime('%H:%M:%S', time.gmtime(riders[i].timegap - leadersTime))))
        pass
    pass

def SaveTotalPoints():

    for i, rider in enumerate(riders):
        print("Total points: " + rider.name + " " + "%0.2f" % (rider.totalPoints,))
        sheet.cell(rider.index, 103).value = rider.totalPoints


def RunTour():
    LoadStats()
    HillStage(1)
    HillStage(2)
    SprintStage(3)
    SprintStage(4)
    HillyITT(5)
    SprintStage(6)
    HillStage(7)
    DownHill(8)
    MTF(9)
    SprintStage(10)
    DownHill(11)
    SprintStage(12)
    SprintStage(13)
    HillStage(14)
    DownHill(15)
    DownHill(16)
    MTF(17)
    MTF(18)
    SprintStage(19)
    HillyITT(20)
    SprintStage(21)

    SaveTotalPoints()
    selectedRidersWorkbook.save("SelectedRiders.xlsx")


if __name__ == '__main__':
    RunTour()
