#Scrapes and appends rider category and cost to the entries

import requests
from bs4 import BeautifulSoup
import rider
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
from unidecode import unidecode

## Scrape pcs, this is necessary because they have the last names in bold, something that Velogames lacks
def ScrapePCS():
    print("Scraping PCS")

    output = open('SelectedRiders.txt', 'w+')
    URL  = "https://www.procyclingstats.com/race/tour-de-france/startlist"
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')

    teams = soup.find_all(class_='team')

    for team in teams:
        results = team.find_all(class_='blue')
        for result in results:
            output.write(unidecode(result.text))
            output.write('\n')
    output.close()

# Scrape Velogames, because that's the game we're playing


    #print(names[i])

#print(str(sheet.max_row))


def ScrapeVelogames():
    workbook = load_workbook(filename="AllCyclists.xlsx")
    sheet = workbook["DYN_cyclist"]

    names = []

    for i in range(sheet.max_row - 1):
        name = str(sheet.cell(i+2,3).value) + " " + str(sheet.cell(i+2,2).value)
        name = unidecode(name).lower()
        names.append(name)

    print("Scraping Velogames")
    URL = 'https://www.velogames.com/velogame/2021/riders.php'
    page = requests.get(URL)

    soup = BeautifulSoup(page.content, 'html.parser')


    rows = soup.find('tbody').find_all("tr")
    for row in rows:
        tds = row.find_all("td")
        # 1 = name
        # 2 = team
        # 3 = category
        # 4 = cost

        #get the first name and last name together as before by stepping through the workbook
        #when this matches that value, get the index, and add the relevant info to the sheet

        thisNam = tds[1].text
        team = tds[2].text
        cat = tds[3].text
        cost = tds[4].text
        thisNam = unidecode(thisNam).lower()

        if names.__contains__(thisNam):
            row = names.index(thisNam) + 2
            #print("attempting " + thisNam + " " + str(row))
            #print(thisNam + " " + team + " " + cat+ " " +cost)

            sheet.cell(row, 100).value = cat
            sheet.cell(row, 101).value = int(cost)
            sheet.cell(row, 102).value = team


        else:
            print(thisNam + " NOT FOUND! (You may have to check how the name is written in AllCyclists.xlxs vs the website)")
            pass

    workbook.save("SelectedRiders.xlsx")

if __name__ == '__main__':
    ScrapePCS()
    ScrapeVelogames()
