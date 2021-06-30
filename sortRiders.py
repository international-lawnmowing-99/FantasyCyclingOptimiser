import rider
from bs4 import BeautifulSoup
import requests

from openpyxl import load_workbook
import NameAngliciser

from openpyxl.worksheet import worksheet
workbook = load_workbook(filename="cyclists.xlsx")
sheet = workbook.active

names = []

for i in range(sheet.max_row - 1):
    nam = sheet.cell(i+2,2).value + " " + sheet.cell(i+2,1).value
    nam = deaccent(nam)
    names.append(nam)
    print(names[i])

print(str(sheet.max_row))

URL = 'https://www.velogames.com/velogame/2021/riders.php'
page = requests.get(URL)

soup = BeautifulSoup(page.content, 'html.parser')


rows = soup.find('tbody').find_all("tr")
for row in rows:
    tds = row.find_all("td")
    # 1 = name
    # 2 = team
    # 3 = category
    # 4 = cost

    #get the first name and last name together as before by stepping through the workbook
    #when this matches that value, get the index, and add the relevant info to the sheet

    thisNam = tds[1].text
    team = tds[2].text
    cat = tds[3].text
    cost = tds[4].text
    thisNam = deaccent(thisNam)
    print(thisNam + " " + team + " " + cat+ " " +cost)
    if names.__contains__(thisNam):
        row = names.index(thisNam) + 2
        print("attempting " + thisNam + " " + str(row))
        sheet.cell(row, 21).value = cat
        sheet.cell(row, 20).value = cost
        sheet.cell(row, 22).value = team


    else:
        print(thisNam + " NOT FOUND")

#workbook.save("theChosenOnes2.xlsx")
