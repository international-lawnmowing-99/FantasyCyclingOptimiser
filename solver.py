from pulp import *
import openpyxl

sheet = openpyxl.load_workbook("SelectedRiders.xlsx", data_only=True)['DYN_cyclist']

model = LpProblem("Cycling_Optimiser", LpMaximize)

function = ""
costConstraint = ""
allRounderConstraint = ""
sprinterContraint = ""
climberConstraint = ""
unclassedConstrant = ""
numberOfRidersConstraint = ""
variables = []

for i in range(184):
    variable = LpVariable(str(i), 0,1,LpInteger)
    variables.append(variable)

    function +=  sheet.cell(i+2, 103).value * variable
    costConstraint += int(sheet.cell(i+2, 101).value) * variable
    numberOfRidersConstraint += variable

    if sheet.cell(i+2, 100).value == "All Rounder":
        allRounderConstraint += variable
    elif sheet.cell(i+2, 100).value == "Climber":
        climberConstraint += variable
    elif sheet.cell(i+2, 100).value == "Unclassed":
        unclassedConstrant += variable
    elif sheet.cell(i+2, 100).value == "Sprinter":
        sprinterContraint += variable
    else:
        print( "error! " + str(sheet.cell(i+2, 2).value) + " has no class! " + str(sheet.cell(i+2, 100).value))

model += function

model += (costConstraint <= 100)
model += (numberOfRidersConstraint == 9)
model += (sprinterContraint >= 1)
model += (climberConstraint >= 2)
model += (unclassedConstrant >= 3)
model += (allRounderConstraint >= 2)


print(type(function))



optimization_result = model.solve()

assert optimization_result == LpStatusOptimal
print("Status:", LpStatus[model.status])
print("Optimal Solution to the problem: ", value(model.objective))
print ("Individual decision_variables: ")
for v in model.variables():
    if v.varValue == 1:
        print(sheet.cell(int(v.name) + 2, 2).value, " ", sheet.cell(int(v.name) + 2, 3).value, " ", sheet.cell(int(v.name) + 2, 100).value)
